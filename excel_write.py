"""
 @Coding: utf-8
 @Product: financial-huanqiu
 @Author: rtf
 @Time: 2019-01-08 16:02
 @FileName: excel_write.py
 @Software: PyCharm Community Edition
"""

import os
import openpyxl
from openpyxl import load_workbook


"""
TEST_CASE_DIR: 测试用例目录
TEST_REPORT_DIR：测试报告目录
excel_path1：测试用例文件
excel_path2：测试报告文件
"""


CUR_FILE_PATH = os.path.dirname(os.path.realpath(__file__))
TEST_CASE_DIR = os.path.join(os.path.dirname(CUR_FILE_PATH), "test_ddt/test_case")
TEST_REPORT_DIR = os.path.join(os.path.dirname(CUR_FILE_PATH), "test_ddt/report")


def copy_excel(excel_path1, excel_path2):
    excel_path1 = f"{TEST_CASE_DIR}/{excel_path1}"
    excel_path2 = f"{TEST_REPORT_DIR}/{excel_path2}"

    # copy excel，excel_path1 copy to excel_path2
    wb2 = openpyxl.Workbook()
    wb2.save(excel_path2)

    # read excel data
    wb1 = openpyxl.load_workbook(excel_path1)
    wb2 = openpyxl.load_workbook(excel_path2)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]
    max_row = sheet1.max_row
    max_column = sheet1.max_column

    for m in list(range(1,max_row+1)):
        for n in list(range(97,97+max_column)):   # chr(97)='a'
            n = chr(n)                            # ASCII字符
            i ='%s%d'% (n, m)                     # 单元格编号
            cell1 = sheet1[i].value
            sheet2[i].value = cell1
    wb2.save(excel_path2)
    wb1.close()
    wb2.close()


class WriteExcel(object):
    # modify excel data
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = load_workbook(self.file_name)
        self.ws = self.wb.active

    def write(self, row_n, col_n, value):
        # 写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.file_name)

if __name__ == "__main__":
    copy_excel("case_user.xlsx", "case_user_report.xlsx")

    # wt = WriteExcel("case_user_report.xlsx")
    # wt.write(4, 5, "HELLEOP")
    # wt.write(4, 6, "HELLEOP")